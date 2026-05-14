import streamlit as st
import pandas as pd
import plotly.express as px
import os
import base64
import io
from datetime import datetime, timedelta
from github import Github
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- 0. CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Inteligência de Giro - Tecadi", layout="wide", initial_sidebar_state="expanded")

def aplicar_estilo_blindado():
    BG_PATH = "assets/tecadi.png"
    estilo = """
    <style>
    .element-container:has(style) { display: none; }
    [data-testid="stSidebar"] { background-color: #003366 !important; }
    [data-testid="stSidebar"] h3, [data-testid="stSidebar"] label, 
    [data-testid="stSidebar"] p, [data-testid="stSidebar"] .stMarkdownContainer p {
        color: #FFFFFF !important; font-weight: 600 !important;
    }
    [data-testid="stFileUploadDropzone"] span, [data-testid="stFileUploadDropzone"] small,
    [data-testid="stFileUploadDropzone"] p, [data-testid="stFileUploadDropzone"] div {
        color: #000000 !important;
    }
    [data-testid="stSidebar"] input, [data-testid="stSidebar"] div[data-baseweb="select"] span {
        color: #000000 !important;
    }
    .stApp { color: #1E1E1E; }
    h1, h2, h3 { color: #003366 !important; font-weight: bold !important; }
    [data-testid="stMetricLabel"] * { color: #000000 !important; font-weight: bold !important; font-size: 16px !important;}
    [data-testid="stMetricValue"] { color: #003366 !important; font-weight: 900 !important;}
    </style>
    """
    st.markdown(estilo, unsafe_allow_html=True)
    if os.path.exists(BG_PATH):
        with open(BG_PATH, "rb") as f: data = base64.b64encode(f.read()).decode()
        st.markdown(f"""<style>.stApp {{ background-image: url("data:image/png;base64,{data}"); background-size: cover; background-attachment: fixed; }}</style>""", unsafe_allow_html=True)

aplicar_estilo_blindado()

# --- 1. MOTORES DE DADOS E INTELIGÊNCIA ---
HISTORICO_PATH = "historico_seriais.parquet"
ENDERECOS_PATH = "Endereços Transitorios.xlsx"

@st.cache_data(ttl=60)
def get_enderecos_vivo():
    """Lê a planilha de regras de negócio (SLA) de forma inteligente à prova de erros de digitação."""
    if not os.path.exists(ENDERECOS_PATH): return pd.DataFrame()
    df_e = pd.read_excel(ENDERECOS_PATH)
    df_e.rename(columns=lambda x: str(x).strip(), inplace=True)
    
    col_end, col_az, col_resp, col_rota = None, None, None, None
    for c in df_e.columns:
        if 'endere' in c.lower() or 'endereço' in c.lower(): col_end = c
        elif c.lower() == 'az': col_az = c
        elif 'respons' in c.lower(): col_resp = c
        elif 'rota' in c.lower() or 'tempo' in c.lower() or 'dia' in c.lower(): col_rota = c

    rename_dict = {}
    if col_end: rename_dict[col_end] = 'ENDEREÇO_REF'
    if col_az: rename_dict[col_az] = 'AZ_REF'
    if col_resp: rename_dict[col_resp] = 'RESPONSAVEL_REF'
    if col_rota: rename_dict[col_rota] = 'Limite_SLA_Dias'
    
    df_e.rename(columns=rename_dict, inplace=True)
    return df_e

@st.cache_data(ttl=60)
def load_data():
    if os.path.exists(HISTORICO_PATH):
        df = pd.read_parquet(HISTORICO_PATH)
        df['DT Entrada'] = pd.to_datetime(df['DT Entrada'], errors='coerce')
        if 'DT Ultimo Endereco' not in df.columns:
            df['DT Ultimo Endereco'] = df['DT Entrada']
        else:
            df['DT Ultimo Endereco'] = pd.to_datetime(df['DT Ultimo Endereco'], errors='coerce')
        if 'Dias End. Atual' not in df.columns:
            df['Dias End. Atual'] = df['Dias Pendentes'] if 'Dias Pendentes' in df.columns else 0
        return df
    return pd.DataFrame()

def push_to_github():
    try:
        g = Github(st.secrets["GITHUB_TOKEN"])
        repo = g.get_repo(st.secrets["REPO_NAME"])
        with open(HISTORICO_PATH, "rb") as f: content = f.read()
        try:
            contents = repo.get_contents(HISTORICO_PATH, ref="main")
            repo.update_file(contents.path, f"Sinc: {datetime.now().strftime('%d/%m/%Y %H:%M')}", content, contents.sha, branch="main")
        except:
            repo.create_file(HISTORICO_PATH, "Base Inicial", content, branch="main")
        return True
    except Exception as e:
        st.error(f"Erro GitHub: {e}")
        return False

def processar_motor(arquivo_novo, data_selecionada):
    df_batimento = pd.read_excel(arquivo_novo, skiprows=6)
    df_enderecos_vivo = get_enderecos_vivo()
    
    if df_enderecos_vivo.empty:
        st.error("Erro: A planilha Endereços Transitorios.xlsx não foi encontrada!")
        return False
        
    df_batimento.rename(columns={'Data doc': 'DT Doc', 'Data Serial': 'DT Serial'}, inplace=True)
    
    # Faz o filtro de segurança na porta
    df_atual = pd.merge(df_batimento, df_enderecos_vivo, left_on='Endereço', right_on='ENDEREÇO_REF', how='inner')
    data_operacao = pd.to_datetime(data_selecionada)
    df_hist = load_data()

    if df_hist.empty:
        df_hist = df_atual.copy()
        df_hist['DT Entrada'] = data_operacao
        df_hist['DT Ultimo Endereco'] = data_operacao
        df_hist['Endereço Anterior'] = "-"
        df_hist['Qtd Movimentações'] = 0
        df_hist['Status'] = 'Em Trânsito'
        df_hist['Dias Pendentes'] = 0
        df_hist['Dias End. Atual'] = 0
    else:
        df_hist_ativo = df_hist[df_hist['Status'] == 'Em Trânsito'].copy()
        ativos_lista = df_hist_ativo['Serial'].tolist()
        
        df_mantidos = df_atual[df_atual['Serial'].isin(ativos_lista)].copy()
        old_end = df_hist_ativo.set_index('Serial')['Endereço']
        old_mov = df_hist_ativo.set_index('Serial')['Qtd Movimentações']
        old_dt = df_hist_ativo.set_index('Serial')['DT Entrada']
        old_ant = df_hist_ativo.set_index('Serial')['Endereço Anterior']
        old_dt_ult = df_hist_ativo.set_index('Serial')['DT Ultimo Endereco']
        
        df_mantidos['DT Entrada'] = df_mantidos['Serial'].map(old_dt)
        df_mantidos['Endereço Antigo Memoria'] = df_mantidos['Serial'].map(old_end)
        df_mantidos['DT Ultimo Endereco'] = df_mantidos['Serial'].map(old_dt_ult)
        
        mudou_mask = df_mantidos['Endereço'] != df_mantidos['Endereço Antigo Memoria']
        df_mantidos['Endereço Anterior'] = df_mantidos['Serial'].map(old_ant)
        df_mantidos.loc[mudou_mask, 'Endereço Anterior'] = df_mantidos.loc[mudou_mask, 'Endereço Antigo Memoria']
        df_mantidos.loc[mudou_mask, 'DT Ultimo Endereco'] = data_operacao 
        df_mantidos['Qtd Movimentações'] = df_mantidos['Serial'].map(old_mov).fillna(0)
        df_mantidos.loc[mudou_mask, 'Qtd Movimentações'] += 1
        
        df_mantidos['Status'] = 'Em Trânsito'
        df_mantidos['Dias Pendentes'] = (data_operacao - pd.to_datetime(df_mantidos['DT Entrada'])).dt.days
        df_mantidos['Dias End. Atual'] = (data_operacao - pd.to_datetime(df_mantidos['DT Ultimo Endereco'])).dt.days
        df_mantidos = df_mantidos.drop(columns=['Endereço Antigo Memoria'])
        
        df_novos = df_atual[~df_atual['Serial'].isin(ativos_lista)].copy()
        df_novos['DT Entrada'] = data_operacao
        df_novos['DT Ultimo Endereco'] = data_operacao
        df_novos['Endereço Anterior'] = "-"
        df_novos['Qtd Movimentações'] = 0
        df_novos['Status'] = 'Em Trânsito'
        df_novos['Dias Pendentes'] = 0
        df_novos['Dias End. Atual'] = 0
        
        df_saidas = df_hist_ativo[~df_hist_ativo['Serial'].isin(df_atual['Serial'].tolist())].copy()
        df_saidas['Status'] = 'Finalizado'
        df_saidas['DT Saída'] = data_operacao
        
        df_hist = pd.concat([df_hist[df_hist['Status'] == 'Finalizado'], df_saidas, df_mantidos, df_novos], ignore_index=True)

    # Limpeza de colunas temporárias de SLA antes de salvar (mantém o banco leve)
    cols_to_drop = [c for c in ['ENDEREÇO_REF', 'AZ_REF', 'RESPONSAVEL_REF', 'Limite_SLA_Dias'] if c in df_hist.columns]
    df_hist.drop(columns=cols_to_drop, inplace=True, errors='ignore')
    
    df_hist.to_parquet(HISTORICO_PATH, index=False)
    return push_to_github()

# --- 2. SIDEBAR ---
with st.sidebar:
    LOGO_PATH = "assets/logosemfundotecadi.png"
    if os.path.exists(LOGO_PATH): st.image(LOGO_PATH, use_container_width=True)
    
    st.markdown("### 🔄 Atualização de Dados")
    df_full = load_data()
    ultima_data_banco = None
    
    if not df_full.empty:
        df_pendente_temp = df_full[df_full['Status'] == 'Em Trânsito']
        if not df_pendente_temp.empty:
            ultima_data_banco = (df_pendente_temp['DT Entrada'] + pd.to_timedelta(df_pendente_temp['Dias Pendentes'], unit='D')).max()
        else:
            ultima_data_banco = pd.to_datetime(df_full['DT Saída'], errors='coerce').max()
        st.warning(f"📌 **Última base lida: {ultima_data_banco.strftime('%d/%m/%Y')}**")

    data_batimento = st.date_input("Data do Novo Relatório:", value=datetime.today(), format="DD/MM/YYYY")
    arquivo = st.file_uploader("Anexe o Batimento (.xlsx)", type=['xlsx'])
    
    data_ja_processada = False
    if ultima_data_banco and pd.to_datetime(data_batimento) <= ultima_data_banco:
        st.error("⚠️ Esta data já consta no sistema ou é anterior à última atualização.")
        data_ja_processada = True

    if arquivo:
        btn_label = "Sincronizar Retenção" if not data_ja_processada else "Sobrescrever Dados (Cuidado)"
        btn_type = "primary" if not data_ja_processada else "secondary"
        if st.button(btn_label, use_container_width=True, type=btn_type):
            with st.spinner("Mapeando histórico e atualizando SLAs..."):
                if processar_motor(arquivo, data_batimento):
                    st.success("Sincronizado com Sucesso!")
                    st.rerun()

    st.markdown("---")
    
    df_pendente = df_full[df_full['Status'] == 'Em Trânsito'].copy() if not df_full.empty else pd.DataFrame()

    # 🔥 MÁGICA: APLICAÇÃO DINÂMICA DO SLA EM TEMPO REAL 🔥
    if not df_pendente.empty:
        df_end_live = get_enderecos_vivo()
        if not df_end_live.empty:
            df_pendente = pd.merge(df_pendente, df_end_live, left_on='Endereço', right_on='ENDEREÇO_REF', how='left')
            
            # Sobrescreve AZ e RESPONSAVEL vivos do Excel
            if 'AZ_REF' in df_pendente.columns: df_pendente['AZ'] = df_pendente['AZ_REF'].fillna(df_pendente.get('AZ', '-'))
            if 'RESPONSAVEL_REF' in df_pendente.columns: df_pendente['RESPONSAVEL'] = df_pendente['RESPONSAVEL_REF'].fillna(df_pendente.get('RESPONSAVEL', '-'))
            if 'Limite_SLA_Dias' not in df_pendente.columns: df_pendente['Limite_SLA_Dias'] = 1
            
            # Função implacável de Prioridade
            def calc_pri_vivo(row):
                lim = row.get('Limite_SLA_Dias', 1)
                d = row.get('Dias End. Atual', 0)
                try: lim = int(lim)
                except: lim = 1
                try: d = int(d)
                except: d = 0
                
                if d > lim: return 'ESTOURADO'
                elif d == lim or (lim - d) <= 1: return 'CRÍTICO'
                else: return 'Normal'
                
            df_pendente['Prioridade'] = df_pendente.apply(calc_pri_vivo, axis=1)

    if not df_pendente.empty:
        st.markdown("### 🔍 Filtros")
        f_az = st.selectbox("Filtrar por AZ", ["Todos"] + sorted(df_pendente['AZ'].dropna().unique().tolist()))
        if f_az != "Todos": df_pendente = df_pendente[df_pendente['AZ'] == f_az]

# --- 3. DASHBOARD ---
st.title("📦 Hub de Retenção Transitória")

if df_full.empty or df_pendente.empty:
    st.info("A base está vazia ou não há seriais no transitório.")
    st.stop()

c1, c2, c3, c4 = st.columns(4)
with c1: st.metric("Seriais no Transitório", len(df_pendente))

qtd_estourados = len(df_pendente[df_pendente['Prioridade'] == 'ESTOURADO']) if 'Prioridade' in df_pendente.columns else 0
qtd_criticos = len(df_pendente[df_pendente['Prioridade'] == 'CRÍTICO']) if 'Prioridade' in df_pendente.columns else 0

with c2: st.metric("🚨 SLA Estourado", qtd_estourados)
with c3: st.metric("⚠️ Risco Crítico", qtd_criticos)
with c4: st.metric("🔄 Mov. Internas", len(df_pendente[df_pendente['Qtd Movimentações'] > 0]) if 'Qtd Movimentações' in df_pendente.columns else 0)

st.markdown("---")

st.subheader("🔥 Foco de Atuação: SLAs Estourados")
if 'Prioridade' in df_pendente.columns:
    df_problema = df_pendente[df_pendente['Prioridade'] == 'ESTOURADO']
    if not df_problema.empty:
        grafico_problema = df_problema.groupby(['AZ', 'RESPONSAVEL']).size().reset_index(name='Qtd_Retida')
        fig = px.bar(grafico_problema, x="AZ", y="Qtd_Retida", color="RESPONSAVEL", text_auto=True, template="plotly_white", color_discrete_sequence=px.colors.qualitative.Bold)
        fig.update_layout(xaxis={'categoryorder':'total descending'}, legend_title_text="Dono da Área")
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.success("🎉 Nenhum serial estourou o limite de seus respectivos endereços.")

st.markdown("---")
st.subheader("📋 Mapa de Rastreio Operacional")

colunas_visoes = ['Serial', 'AZ', 'RESPONSAVEL', 'Prioridade', 'Dias End. Atual', 'Limite_SLA_Dias', 'Dias Pendentes', 'Qtd Movimentações', 'Endereço', 'Endereço Anterior', 'DT Entrada']
colunas_existentes = [col for col in colunas_visoes if col in df_pendente.columns]

df_detalhe = df_pendente[colunas_existentes].copy()

renomeacoes = {
    'Endereço': 'Endereço Atual', 
    'Endereço Anterior': 'End. Anterior', 
    'Qtd Movimentações': 'Mov. Internas', 
    'DT Entrada': 'DT Chegada (Inicial)',
    'Limite_SLA_Dias': '(Dias Permitidos no Endereço)',
    'Dias Pendentes': 'Dias Totais (Trânsito)'
}

df_detalhe.rename(columns=renomeacoes, inplace=True)

if 'DT Chegada (Inicial)' in df_detalhe.columns:
    df_detalhe['DT Chegada (Inicial)'] = df_detalhe['DT Chegada (Inicial)'].dt.strftime('%d/%m/%Y')

colunas_ordenacao = []
if 'Prioridade' in df_detalhe.columns: colunas_ordenacao.append('Prioridade')
if 'Dias End. Atual' in df_detalhe.columns: colunas_ordenacao.append('Dias End. Atual')

if colunas_ordenacao:
    # Ajuste de peso para ordenar ESTOURADO > CRÍTICO > Normal
    ordem_pri = {'ESTOURADO': 1, 'CRÍTICO': 2, 'Normal': 3}
    if 'Prioridade' in df_detalhe.columns:
        df_detalhe['Ordem_Temp'] = df_detalhe['Prioridade'].map(ordem_pri).fillna(4)
        df_detalhe = df_detalhe.sort_values(by=['Ordem_Temp', 'Dias End. Atual'], ascending=[True, False]).drop(columns=['Ordem_Temp'])

st.dataframe(df_detalhe, use_container_width=True, hide_index=True)

# --- EXPORTAÇÃO EXCEL ---
buffer = io.BytesIO()
with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
    df_detalhe.to_excel(writer, index=False, sheet_name='Analise_Retencao')
    workbook = writer.book
    worksheet = writer.sheets['Analise_Retencao']
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in worksheet.iter_cols(min_row=1, max_row=1, max_col=worksheet.max_column):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
    for col in worksheet.iter_cols(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in col:
            cell.alignment = center_alignment
            cell.border = thin_border
            
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = length + 4

st.download_button(
    label="📥 Extrair Relatório em Excel",
    data=buffer.getvalue(),
    file_name=f"Analise_Giro_Tecadi_{datetime.now().strftime('%d%m%Y')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary"
)
